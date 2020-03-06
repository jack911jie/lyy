import os
from bs4 import BeautifulSoup
import requests
import re
import xlrd
from openpyxl import Workbook
import time
from tqdm import tqdm

def date_time():
    dt=time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()) )
    return dt

def main_spider():
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36'}
    #payload = {'spm':'a213w.3064813.a214dqe.22','province': '%B9%E3%CE%F7'}
    prov= r'?province=%B9%E3%CE%F7&page=1'
    try:
        req = requests.get('https://sf.taobao.com/item_list.htm'+prov,timeout=5)
       # req = requests.get('http://www.lanbama.cn/test.html', timeout=5)
       # print('正在连接',req.request.url)
        print('正在连接淘宝司法拍卖网页_广西')
        soup = BeautifulSoup(req.content, "lxml")
        c = soup.get_text()
        k = int(soup.find("em", class_='page-total').text)
        total_count=int(soup.find("em", class_='count').text)
        result = re.compile(r'\"title\"\:\"(.*?)\"\,\"picUrl\"').findall(c)
        print('找到数据 ', k, ' 页，共 ', total_count, '条。')

        page_count=[]
        for i in range(k):
            page_count.append(i+1)

        #m = [1,2]
        prgs=''
        #for i in tqdm(m, ncols=75):
        for i in tqdm(page_count,ncols=75):
            out = []
            provB = r'province=%B9%E3%CE%F7&page=' + str(i)
            rB = requests.get("https://sf.taobao.com/item_list.htm", params=provB)
            #rB = requests.get('http://www.lanbama.cn/test.html')
            # print(rB.status_code, "|", rB.request.url)
            #print('找到共 ',k,' 页，共 ',total_count,'条数据，正在获取第 ',i,' 页的数据。')
            prgs = prgs + str(i)
            time.sleep(0.2)
            soupB = BeautifulSoup(rB.content, "lxml")
            cB = soupB.get_text()
            out = re.compile(r'\"title\"\:\"(.*?)\"\,\"picUrl\"').findall(cB)
            result.extend(out)
    except:
        print("网络连接失败")
        result=[]
    else:
        print('数据下载及提取完毕，准备开始比对。')

    return result

def read_excel():
    file = os.getcwd() + "\\list.xlsx"
    if os.path.isfile(file):
        wb = xlrd.open_workbook(filename=file)  # 打开文件
        sht=wb.sheet_by_index(0)
        lis=sht.col_values(1,2)

        nrows = sht.nrows
        ncols = sht.ncols

        lis = []
        for i in range(nrows - 2):
            lis_r = []
            for j in range(2):
                lis_r.append(sht.cell(i + 2, j).value)
            lis.append(lis_r)
        return lis
    else:
        print("文件list.xlsx不存在。请检查。")
        input("")
        os._exit(0)

def toMatch(local,remote):
    ret = list(set(local).intersection(set(remote)))
    c=len(ret)
    return ret

def toMatch_fuzzy(local,remote):
    print('正在比对数据')
    res=[]
    prgs=""
    if remote:
        for i in tqdm(local, ncols=75):
            prgs=prgs+i[1]
            time.sleep(0.2)
            res_r=[]
            for j in remote:
                if i[1] in j:
                    res_r.append(j)
                    if i[0]:
                        res_r.append(i[0])
                    else:
                        res_r.append("-")
                    res.append(res_r)
        print('数据比对结束，共找到 ',len(res),'条可能相同的数据。')
        return res
    else:
        print('无网络数据，无法比对。')
        print("——End——")
    return res

def write_excel(list):
    if list:
        print('正在写入数据')

        wb = Workbook()
        ws=wb.active
        ws['A1']='案件名称'
        ws['B1']='匹配信息'
        for i in range(0,len(list)):
            ws.cell(row=i+2,column=1,value=list[i][1])
            ws.cell(row=i + 2, column=2, value=list[i][0])

        path = os.getcwd()+"\\"
        pre=date_time()
        filename = path+pre+".xlsx"
        wb.save(filename)
        print('写入 '+str(len(list))+' 条数据')
        print("——End——")
    else:
        print("")

if __name__=="__main__":
    local=read_excel()
    remote=main_spider()
    result=toMatch_fuzzy(local,remote)
    write_excel(result)
    #print(result)