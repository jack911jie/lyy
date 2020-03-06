import os
import random
from bs4 import BeautifulSoup
import requests
import re
import xlrd
from openpyxl import Workbook
import time
from tqdm import tqdm

def date_time():
    dt=time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()) )
    return dt

def spider_zwr():
    try:
        print('正在连接【全国企业破产重整案件信息网】……')
        d = {"zcd":[999],"hy":[999],"gsxz":[999],"zczb":[999],"zgsl":[999],"qylx":[999],"ajzt":[999],\
             "sfzkzmtzr":"2","zmjzsjpx":"2","slfy":"","glr":"","pageNum":1,"sjlx":1,"qymc": ""}
        req = requests.post('http://pccz.court.gov.cn/pcajxxw/pctzr/tzrlb', data=d, timeout=5)
        soup = BeautifulSoup(req.content, "lxml")
        #print(soup)
        totalNum = int(re.compile(r'共<span>(.*?)</span>位债务人').findall(str(soup))[0])

        if totalNum % 10 != 0:
            totalPage = totalNum // 10 + 1
        else:
            totalPage = totalNum / 10
        print('找到', totalNum, '条数据，共', totalPage, '页。')

        page_count = []
        for i in range(1, int(totalPage)+1):
            page_count.append(i)

        out = []
        for i in tqdm(page_count, ncols=75):
            d_2 = {"zcd":[999],"hy":[999],"gsxz":[999],"zczb":[999],"zgsl":[999],"qylx":[999],"ajzt":[999],\
             "sfzkzmtzr":"2","zmjzsjpx":"2","slfy":"","glr":"","pageNum":int(i),"sjlx":1,"qymc": ""}
            req_2 = requests.post('http://pccz.court.gov.cn/pcajxxw/pctzr/tzrlb', data=d_2, timeout=5)
            time.sleep(random.randint(0, 3))
            soup_2 = BeautifulSoup(req_2.content, "lxml")
            c_2 = soup_2.find_all("a")
            PageContent = re.compile(r'title="(.*?)\">').findall(str(c_2))
            if PageContent:
                out.extend(PageContent)

        # n = 1
        # for k in out:
        #     print(n, ":", k)
        #     n += 1
    except:
        print("网络连接失败")
        out = []
    else:
        print('数据下载及提取完毕，准备开始比对。')
    return out

def spider_pcgg():
    try:
        print('正在连接【全国企业破产重整案件信息网】……')
        d = {'start': '', 'end': '', 'cbt': '', 'wslx': 0, 'ajlx': '999', 'pageNum': 1}
        req = requests.post('http://pccz.court.gov.cn/pcajxxw/pcgg/gglb', data=d, timeout=5)
        soup = BeautifulSoup(req.content, "lxml")
        totalNum = int(re.compile(r'共计<span>(.*?)</span>件').findall(str(soup))[0])
        if totalNum % 10 != 0:
            totalPage = totalNum // 10 + 1
        else:
            totalPage = totalNum / 10
        print('找到',totalNum,'条数据，共',totalPage,'页。')
        c = soup.find_all("a")
        FirstPageContent = re.compile(r'title="(.*?)\">').findall(str(c))
        out = []
        if FirstPageContent:
            out = FirstPageContent

        page_count = []
        for i in range(2,4):
            page_count.append(i)

        for i in tqdm(page_count,ncols=75):
            d_2 = {'start': '', 'end': '', 'cbt': '', 'wslx': 0, 'ajlx': '999', 'pageNum': int(i)}
            req_2 = requests.post('http://pccz.court.gov.cn/pcajxxw/pcgg/gglb', data=d_2, timeout=5)
            soup_2 = BeautifulSoup(req_2.content, "lxml")
            c_2 = soup_2.find_all("a")
            PageContent = re.compile(r'title="(.*?)\">').findall(str(c_2))
            out.extend(PageContent)

        n = 1
        for k in out:
            print(n, ":", k)
            n += 1
    except:
        print("网络连接失败")
        out=[]
    else:
        print('数据下载及提取完毕，准备开始比对。')
    return out

def read_excel():
    file = os.getcwd() + "\\list_pcgg.xlsx"
    if os.path.isfile(file):
        wb = xlrd.open_workbook(filename=file)  # 打开文件
        sht=wb.sheet_by_index(0)
        lis=sht.col_values(1,2)

        nrows = sht.nrows
        ncols = sht.ncols

        lis = []
        for i in range(nrows - 2):
            lis_r = []
            for j in range(2):
                lis_r.append(sht.cell(i + 2, j).value)
            lis.append(lis_r)
        return lis
    else:
        print("文件list_pcgg.xlsx不存在。请检查。")
        input("")
        os._exit(0)

def toMatch(local,remote):
    ret = list(set(local).intersection(set(remote)))
    c=len(ret)
    return ret

def toMatch_fuzzy(local,remote):
    print('正在比对数据')
    res=[]
    prgs=""
    if remote:
        for i in tqdm(local, ncols=75):
            prgs=prgs+i[1]
            time.sleep(0.2)
            res_r=[]
            for j in remote:
                if i[1] in j:
                    res_r.append(j)
                    if i[0]:
                        res_r.append(i[0])
                    else:
                        res_r.append("-")
                    res.append(res_r)
        print('数据比对结束，共找到 ',len(res),'条可能相同的数据。')
        return res
    else:
        print('无网络数据，无法比对。')
        print("——End——")
    return res

def write_excel(list):
    if list:
        print('正在写入数据')

        wb = Workbook()
        ws=wb.active
        ws['A1']='案件名称'
        ws['B1']='匹配信息'
        for i in range(0,len(list)):
            ws.cell(row=i+2,column=1,value=list[i][1])
            ws.cell(row=i + 2, column=2, value=list[i][0])

        path = os.getcwd()+"\\"
        pre="破产管理网比对结果_"+date_time()
        filename = path+pre+".xlsx"
        wb.save(filename)
        print('写入 '+str(len(list))+' 条数据')
        print("——End——")
    else:
        print("")

def write_remote(data):
    filename=os.getcwd()+"\\"+"下载的数据"+date_time()+".txt"
    fn="下载的数据" + date_time() + ".txt"
    if len(data):
        file=open(filename,"w")
        for i in data:
            file.write(i+"\r")
        #print('下载的数据'+str(len(data))+'条，已保存到\"'+fn+'\"中。')
    else:
        print('未下载到数据')

if __name__=="__main__":
    local=read_excel()
    #remote=spider_pcgg()
    remote = spider_zwr()
    write_remote(remote)
    result=toMatch_fuzzy(local,remote)
    write_excel(result)
    input("已完成，按Enter退出。")
    #print(result)