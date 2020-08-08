#!/usr/bin/env python
# coding: utf-8

# In[4]:


import os
import sys
import datetime
import pandas as pd
from pandas.core.frame import DataFrame
import openpyxl
from openpyxl.styles import PatternFill
from tqdm import tqdm
import time
import logging
import win32com.client as win32
import warnings

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.WARNING, format='%(asctime)s - %(funcName)s-%(lineno)d - %(message)s')
logger = logging.getLogger(__name__)

class ShebaoSalaryTable:
    def __init__(self,fn_sb,fn_slry):
        self.fn_sb=os.path.join(os.getcwd(),fn_sb)
        self.fn_slry=os.path.join(os.getcwd(),fn_slry)
        if not os.path.exists(self.fn_sb):
            print('没有这个csv文件，请检查文件名。')
            sys.exit(0)
        if not os.path.exists(self.fn_slry):
            print('没有这个xlsx文件，请检查文件名。')
            sys.exit(0)
 
        
    def WriteToSalary(self):
        print('正在处理……',end='')
        df_sb=pd.read_csv(self.fn_sb)
        df_slry=pd.read_excel(self.fn_slry,skiprows=4,usecols=('c:g'))
        df_slry.rename(columns={'Unnamed: 2':'姓名','Unnamed: 3':'预发工资'},inplace=True)
        for i in range(2):
            df_slry.drop([len(df_slry)-1],inplace=True)
            
        nameList=df_slry['姓名'].values.tolist()
        
#         nameList=['骆莹莹']
        items=['职工基本养老保险','职工基本医疗保险','失业保险']
    
        logging.info(nameList)
        
        output=[]
        for name in nameList:
            eachPerson=[name]
            for item in items:        
                YingJiao=df_sb.loc[(df_sb['姓名']==name)& (df_sb['缴费对象']=='个人') & (df_sb['险种类型']==item)]
                eachPerson.append(YingJiao['本期应缴'].values.tolist()[0])
            output.append(eachPerson)
#         print(output)
        
        toWriteList=DataFrame(output)   
        r,c=toWriteList.shape
#         toto=toWriteList.drop(0,axis=1)
        logging.info(toWriteList)
        
        toWriteXl=openpyxl.load_workbook(self.fn_slry)
        sht=toWriteXl['资管']   
        
        print('\n')
        for i,rr in enumerate(range(0,r)):
            for cc in range(0,c-1):
                sht.cell(rr+6,5+cc).value=output[rr][cc+1]
#             print("进度:{0}%".format(round((i + 1) * 100 / r)), end="\r")
            progress(int(i*100/(r-1)))
            time.sleep(0.2)

        toWriteXl.save(self.fn_slry[0:-5]+'_自动填报.xlsx')
        
        print('\n完成自动填报。保存文件名为：{}'.format(self.fn_slry[0:-5]+'_自动填报.xlsx'))
    
    def compare(self):
        print('正在比对数据……',end='')
        df_sb=pd.read_csv(self.fn_sb)
        df_slry=pd.read_excel(self.fn_slry,skiprows=4,usecols=('c:g'))
        df_slry.rename(columns={'Unnamed: 2':'姓名','Unnamed: 3':'预发工资','养老保险\n8%':'职工基本养老保险','医疗保险\n2%':'职工基本医疗保险','失业保险\n0.5%':'失业保险'},inplace=True)
        for i in range(2):
            df_slry.drop([len(df_slry)-1],inplace=True)
            
        nameList=df_slry['姓名'].values.tolist()
        
#         nameList=['骆莹莹']
        items=['职工基本养老保险','职工基本医疗保险','失业保险']
    
        dataToCompare=df_slry.values.tolist()
        addr_dif=[]
        for i,vs in enumerate(dataToCompare):
            YangLao_sb=df_sb.loc[(df_sb['姓名']== dataToCompare[i][0]) &                                   (df_sb['险种类型']=='职工基本养老保险')&                                   (df_sb['缴费对象']=='个人')]                                 ['本期应缴'].values.tolist()[0]
             
            YiLiao_sb=df_sb.loc[(df_sb['姓名']== dataToCompare[i][0]) &                                   (df_sb['险种类型']=='职工基本医疗保险')&                                   (df_sb['缴费对象']=='个人')]                                 ['本期应缴'].values.tolist()[0]
            
            ShiYe_sb=df_sb.loc[(df_sb['姓名']== dataToCompare[i][0]) &                                   (df_sb['险种类型']=='失业保险')&                                   (df_sb['缴费对象']=='个人')]                                 ['本期应缴'].values.tolist()[0]
            
            if YangLao_sb!=dataToCompare[i][2]:
                addr_dif.append([i,2,YangLao_sb])
                
            if YiLiao_sb!=dataToCompare[i][3]:
                addr_dif.append([i,3,YiLiao_sb])
                
            if ShiYe_sb!=dataToCompare[i][4]:
                addr_dif.append([i,4,ShiYe_sb])
            
                
            logging.info('{0},{1},{2}'.format(YangLao_sb,YiLiao_sb,ShiYe_sb))
#         logging.info(df_slry.values.tolist())
        logging.info(addr_dif)
    
        toMark=openpyxl.load_workbook(self.fn_slry)
        sht=toMark['资管']   
        
        fillColor=PatternFill('solid',fgColor='ff00ee')
        
        
        if addr_dif:
            N=len(addr_dif)
            print('\n')
            for i,rows in enumerate(addr_dif):
                sht.cell(rows[0]+6,rows[1]+3).value=str(sht.cell(rows[0]+6,rows[1]+3).value)+'<'+str(rows[2])+'>'
                sht.cell(rows[0]+6,rows[1]+3).fill=fillColor
                progress(int(i*100/(N-1)))
                time.sleep(0.2)
                
            toMark.save(self.fn_slry[0:-5]+'_标记不同.xlsx')
            print('\n共发现异常数据{}条，已做颜色标记。保存文件名为：{}'.format(len(addr_dif),self.fn_slry[0:-5]+'_标记不同.xlsx'))
            
        else:
            print('\n未发现不同的数据。')     
    
def XlsToXlsx(fn):  
    print('\nxls是旧版excel文件，正在转换……',end='')
    fname=os.path.join(os.getcwd(),fn)
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(fname)
    wb.SaveAs(fname+'x', FileFormat = 51)              # 转成xlsx格式，路径在原路径下
    wb.Close()                               
    excel.Application.Quit()
    print('转换完成\n')
    
def progress(percent=0, width=30):
    left = width * percent // 100
    right = width - left
    print('\r[', '#' * left, ' ' * right, ']',
          f' {percent:.0f}%',
          sep='', end='', flush=True)
    
if __name__=='__main__':
    while True:
        nameSb=input('一、请输入社保数据文件名（带扩展名.csv,直接回车则按默认文件名。)：')
        nameSlry=input('二、请输入薪酬文件名（带扩展名.xlsx，直接回车则按默认文件名。)：')
        slct=input('\n三、选择功能：1.将社保数据写入薪酬表   2.对比社保数据与薪酬表（直接回车则选择1)')

        mth=datetime.datetime.now().month
        dftSbName=str(mth-1)+'月社保数据.csv'
        dftSlryName=str(mth)+'月薪酬表.xlsx'

        if nameSb=='':
            nameSb=dftSbName
        if nameSlry=='':
            nameSlry=dftSlryName
        else:
            if nameSlry[-3:]=='xls':
                XlsToXlsx(nameSlry)
                nameSlry=nameSlry+'x'
            
        if slct=='':
            slct='1'
            t='操作：1.将社保数据写入薪酬表\n'
        else:
            t='操作：2.对比社保数据与薪酬表\n'

        print('\n默认文件名：{0}， {1}，\n{2}'.format(nameSb,nameSlry,t))
        vrfy=input('确定直接回车，重新选择按1后回车。')

        if vrfy=='':
            break

    sb=ShebaoSalaryTable(nameSb,nameSlry)
    if slct=='1':
        sb.WriteToSalary()
        input('\n回车退出')
    elif slct=='2':
        sb.compare()
        input('\n回车退出')
    else:
        print('\n没有该选项')
        input('\n回车退出')
        sys.exit(0)

